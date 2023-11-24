from PyQt5.QtWidgets import QDialog, QTableWidgetItem, QHeaderView, QMessageBox
from PyQt5.uic import loadUiType
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
import os


Ui_Report_Form, QReport_Form = loadUiType('report.ui')


class Report_Form(QDialog, Ui_Report_Form):

    def __init__(self, request, query, state, name):
        super().__init__()
        self.setupUi(self)
        self.state = state
        self.name = name
        self.ntp_arg(request, query)
        self.json = {"Название НТП": "", "Федеральный округ": "", "Субьект Федерации": "", "Город": "", "Вуз": ""}

        self.save_button.clicked.connect(self.save_excel)
        self.cancel_button.clicked.connect(self.cancel)

    def db_load(self, request, db_name = "SUBDlab.db", *args, **kwargs):
        """ Функция заполнения таблицы НТП проектов.
            args:
            db_name:str - название БД
            tb_name:str - название таблицы
            request:str - запрос"""

        db = QSqlDatabase.addDatabase("QSQLITE")
        db.setDatabaseName(db_name)
        db.open()

        query = QSqlQuery(request)
        column_count = query.record().count()
        self.table.setColumnCount(column_count)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        for col in range(column_count):
            header_item = QTableWidgetItem(query.record().fieldName(col))
            self.table.setHorizontalHeaderItem(col, header_item)

        row = 0
        while query.next():
            self.table.setRowCount(row + 1)
            for col in range(column_count):
                item = QTableWidgetItem(str(query.value(col)))
                self.table.setItem(row, col, item)
            row += 1

        db.close()

    def ntp_arg(self, line,  query):
        line = "WITH tt AS(" + line + ")"
        line += query
        self.db_load(line)

    def cancel(self):
        self.reject()

    def show_warning(self, warning):

        """Вывод предупреждения о пользовательской ошибке
            args:
            warning - предупреждениe"""

        QMessageBox.information(self, 'Файл сохранен', warning)

    def save_excel(self):

        self.json["Название НТП"] = self.state['codprogBox']
        self.json["Федеральный округ"] = self.state['regionBox']
        self.json["Субьект Федерации"] = self.state['oblnameBox']
        self.json["Город"] = self.state['cityBox']
        self.json["Вуз"] = self.state['z2Box']

        form_name = f'Форма {self.name}'
        workbook = Workbook()
        sheet = workbook.active

        sheet['A1'].font = Font(bold=True)
        sheet.cell(row=1, column=1, value=form_name)

        sheet['A3'].font = Font(bold=True)
        sheet.cell(row=3, column=1, value="Выставленный фильтр")

        for line, (name, value) in enumerate(self.json.items()):
            sheet.cell(row=line + 4, column=1, value=name if name else "")
            sheet.cell(row=line + 4, column=2, value=value if name else "")

        for j in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(j)
            sheet.cell(row=10, column=j + 1, value=header_item.text() if header_item else "")

        for i in range(self.table.rowCount()):
            for j in range(self.table.columnCount()):
                item = self.table.item(i, j)
                sheet.cell(row=i + 11, column=j + 1, value=item.text() if item else "")

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        dir = f'Отчетная форма {self.name} {datetime.datetime.now().strftime("%d-%m-%Y %H-%M")}.xlsx'
        workbook.save(f'data\\reports\\{dir}')
        self.show_warning(f"Файл успешно сохранен в директории: \n {os.getcwd()}\\data\\reports\\{dir}")
        self.cancel()

