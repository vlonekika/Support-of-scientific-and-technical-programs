from PyQt5.QtWidgets import QDialog, QMessageBox, QTableWidgetItem, QHeaderView
from PyQt5.uic import loadUiType
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
import os


Ui_Order_Form, QOrder_Form = loadUiType('order_ntp.ui')

class Order_Form(QDialog, Ui_Order_Form):

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.db = "SUBDlab.db"
        self.ffin_browser.setStyleSheet("background-color: lightgray;;")
        self.pfin_browser.setStyleSheet("background-color: lightgray;;")
        self.ffin_pfin_browser.setStyleSheet("background-color: lightgray;;")

        sum_pfin = "SELECT sum(pfin) FROM Ntp_proj"
        self.pfin_browser.setPlainText(self.db_loads(sum_pfin)[0])

        sum_ffin = "SELECT sum(ffin) FROM Ntp_proj"
        self.ffin_browser.setPlainText(self.db_loads(sum_ffin)[0])

        per_ffin_pfin = "SELECT ROUND((CAST(sum(ffin) AS REAL)/(sum(pfin)))*100, 4) FROM Ntp_proj"
        self.ffin_pfin_browser.setPlainText(self.db_loads(per_ffin_pfin)[0])

        self.kvartal_comboBox.addItems(["1 квартал", "2 квартал", "3 квартал", "4 квартал"])
        self.kvartal_changed()
        self.kvartal_comboBox.currentIndexChanged.connect(self.kvartal_changed)

        self.input_sum_edit.setText("0")
        self.per_sum_edit.setText("0")

        self.input_sum_edit.textEdited.connect(self.sum_input)
        self.per_sum_edit.textEdited.connect(self.per_input)
        self.check_table_button.clicked.connect(self.get_fin_table)

        self.save.clicked.connect(self.save_order)
        self.cancel_button.clicked.connect(self.cancel)
        self.save_excel.clicked.connect(self.get_excel)


    def db_loads(self, request):

        db = QSqlDatabase.addDatabase("QSQLITE")
        db.setDatabaseName(self.db)
        db.open()
        query = QSqlQuery(request)
        column_values = []
        while query.next():
            column_value = query.value(0)
            column_values.append(str(column_value))
        db.close()
        return column_values

    def show_warning(self, warning):

        """Вывод предупреждения о пользовательской ошибке
            args:
            warning - предупреждениe"""

        QMessageBox.information(self, 'Предупреждение', warning)

    def kvartal_changed(self):
        self.ffin_kvartal_browser.setStyleSheet("background-color: lightgray;;")
        self.pfin_kvartal_browser.setStyleSheet("background-color: lightgray;;")

        kvartal = self.kvartal_comboBox.currentText()[0]
        self.proj_label.setText(f"""Проект распределения о финансирование НТП в {kvartal} квартале""")

        ffin_kvartal = f"SELECT sum(ffin{kvartal}) FROM Ntp_proj"
        self.ffin_kvartal_browser.setPlainText(self.db_loads(ffin_kvartal)[0])

        pfin_kvartal = f"SELECT sum(pfin{kvartal}) FROM Ntp_proj"
        self.pfin_kvartal_browser.setPlainText(self.db_loads(pfin_kvartal)[0])
        request = f"""WITH tt AS (SELECT ROW_NUMBER() OVER (ORDER BY isp) AS "№", isp, ffin{kvartal}
                             FROM Ntp_proj)
                             SELECT №, isp AS 'ВУЗ', ffin{kvartal} AS 'Финан. в кв. {kvartal}' FROM tt
                             UNION SELECT "Итого", NULL, (SELECT sum(ffin{kvartal}) FROM tt)"""
        self.load_table(request)

    def sum_input(self, value):

        if not value:
            amount = 0
        else:
            try:
                amount = int(value)
                if amount < 0:
                    amount = 0
                    self.show_warning("Введенная сумма должна быть положительной!")
                elif amount + int(self.ffin_kvartal_browser.toPlainText()) > int(self.pfin_browser.toPlainText()):
                    self.input_sum_edit.clear()
                    amount = 0
                    self.show_warning("Введенная сумма не может быть больше суммарного планового финансирования!")
            except ValueError:
                amount = 0
                self.show_warning("Введенная сумма должна быть целым числом!")

        self.input_sum_edit.setText(str(amount))
        line = int(amount)/int(self.pfin_kvartal_browser.toPlainText()) * 100
        self.per_sum_edit.setText(str(round(line, 2)))

    def per_input(self, value):
        if not value:
            amount = 0
        else:

            try:
                amount = int(value)
                amount_sum = (amount * int(self.pfin_kvartal_browser.toPlainText()))/100
                amount_sum += int(self.ffin_kvartal_browser.toPlainText())
                if amount < 0 or amount_sum > int(self.pfin_browser.toPlainText()):
                    self.per_sum_edit.clear()
                    amount = 0
                    warning = "Введенный процент должен быть не меньше 0 и в пересчете не превышать планового финансирования!"
                    self.show_warning(warning)
            except ValueError:
                amount = 0
                self.show_warning("Введенная сумма должна быть целым числом!")

        self.per_sum_edit.setText(str(amount))
        line = int(amount)*int(self.pfin_kvartal_browser.toPlainText())/ 100
        self.input_sum_edit.setText(str(int(line)))

    def load_table(self, request):
        db = QSqlDatabase.addDatabase("QSQLITE")
        db.setDatabaseName(self.db)
        db.open()

        query = QSqlQuery(request)
        column_count = query.record().count()
        self.table.setColumnCount(column_count)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        for col in range(column_count):
            header_item = QTableWidgetItem(query.record().fieldName(col))
            self.table.setHorizontalHeaderItem(col, header_item)

        row = 0
        while query.next():
            self.table.setRowCount(row + 1)
            for col in range(column_count):
                item = QTableWidgetItem(str(query.value(col)))
                self.table.setItem(row, col, item)
            row += 1
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.table.setColumnWidth(0, 60)
        self.table.setColumnWidth(1, 110)
        self.table.setColumnWidth(2, 130)
        db.close()

    def get_fin_table(self):
        per = int(float(self.per_sum_edit.text()))/100
        kvartal = self.kvartal_comboBox.currentText()[0]
        request = f"""WITH tt AS (SELECT ROW_NUMBER() OVER (ORDER BY isp) AS "№", isp, 
                      (ffin{kvartal} + CAST(CAST(pfin{kvartal} AS REAL)*{per} AS INTEGER)) AS ffin
                      FROM Ntp_proj)
                      SELECT №, isp AS 'ВУЗ', ffin AS 'Финан. в кв. {kvartal}' FROM tt
                      UNION SELECT "Итого", NULL, (SELECT sum(ffin) FROM tt)"""
        self.load_table(request)

    def save_order(self):

        per = int(float(self.per_sum_edit.text())) / 100
        kvartal = self.kvartal_comboBox.currentText()[0]

        request = f"""UPDATE ntp_proj
                      SET 
                      ffin{kvartal} = ffin{kvartal}+ CAST(CAST(pfin{kvartal} AS REAL)*{per} AS INTEGER);"""
        self.update_db(request)
        request = f"""UPDATE ntp_proj
                              SET 
                              ffin = ffin1 + ffin2 + ffin3 + ffin4;"""
        self.update_db(request)
        request = """WITH tt AS (SELECT codprog, sum(pfin) AS pfin, sum(pfin1) AS pfin1, sum(pfin2) AS pfin2, 
                                        sum(pfin3) AS pfin3, sum(pfin4) AS pfin4, sum(ffin) AS ffin, 
                                        sum(ffin1) AS ffin1, sum(ffin2) AS ffin2, sum(ffin3) AS ffin3, 
                                        sum(ffin4) AS ffin4 
                     FROM ntp_proj
                     GROUP BY codprog)

                     UPDATE Ntp_prog 
                     SET
                     ffin = (SELECT ffin FROM tt WHERE tt.codprog = Ntp_prog.codprog),
                     ffin1 = (SELECT ffin1 FROM tt WHERE tt.codprog = Ntp_prog.codprog),
                     ffin2 = (SELECT ffin2 FROM tt WHERE tt.codprog = Ntp_prog.codprog),
                     ffin3 = (SELECT ffin3 FROM tt WHERE tt.codprog = Ntp_prog.codprog),
                     ffin4 =  (SELECT ffin4 FROM tt WHERE tt.codprog = Ntp_prog.codprog);"""
        self.update_db(request)
        self.show_warning("Распоряжение о финансирование НТП вынесено!")
        self.get_excel()

    def update_db(self, request):
        db = QSqlDatabase.addDatabase('QSQLITE')
        db.setDatabaseName(self.db)
        db.open()
        query = QSqlQuery()
        query.prepare(request)
        query.exec_()
        db.close()

    def cancel(self):
        self.reject()

    def get_excel(self):

        self.get_fin_table()
        workbook = Workbook()
        sheet = workbook.active

        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)
        sheet['C1'].font = Font(bold=True)
        for j in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(j)
            sheet.cell(row=1, column=j + 1, value=header_item.text() if header_item else "")

        for i in range(self.table.rowCount()):
            for j in range(self.table.columnCount()):
                item = self.table.item(i, j)
                sheet.cell(row=i + 2, column=j + 1, value=item.text() if item else "")

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        dir = f'Распоряжение от {datetime.datetime.now().strftime("%d-%m-%Y %H-%M")}.xlsx'
        workbook.save(f'data\\orders\\{dir}')
        self.show_warning(f"Файл успешно сохранен в директории: \n {os.getcwd()}\\data\\orders\\{dir}")
        self.cancel()

